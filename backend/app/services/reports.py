"""Recruitment reports and their CSV / Excel / PDF exports."""

from __future__ import annotations

import csv
import io
from datetime import UTC, date, datetime

from sqlalchemy.orm import Session

from app.core.constants import CandidateStatus
from app.core.exceptions import ValidationAppError
from app.core.logging import get_logger
from app.repositories import candidate_repository as repo
from app.schemas.dashboard import (
    NamedValue,
    PipelineStage,
    RecruitmentKPIs,
    ReportResponse,
    TrendPoint,
)
from app.services.candidates import STATUS_LABELS
from app.services.matching import analyze_skill_gaps

logger = get_logger(__name__)

PIPELINE_ORDER = (
    CandidateStatus.NEW,
    CandidateStatus.PENDING_REVIEW,
    CandidateStatus.IN_REVIEW,
    CandidateStatus.SHORTLISTED,
    CandidateStatus.INTERVIEWING,
    CandidateStatus.OFFERED,
    CandidateStatus.HIRED,
    CandidateStatus.REJECTED,
    CandidateStatus.ON_HOLD,
)


def build_report(
    session: Session,
    *,
    gap_skills: list[str] | None = None,
    months: int = 6,
    period_start: date | None = None,
    period_end: date | None = None,
) -> ReportResponse:
    processing = repo.resume_processing_stats(session)
    total_candidates = repo.count_candidates(session)
    shortlisted = repo.count_candidates(session, status=CandidateStatus.SHORTLISTED.value)
    rejected = repo.count_candidates(session, status=CandidateStatus.REJECTED.value)

    from sqlalchemy import func, select

    from app.models.ai import MatchRun

    matches_run = session.scalar(select(func.count(MatchRun.id))) or 0

    kpis = RecruitmentKPIs(
        total_candidates=total_candidates,
        resumes_processed=processing["completed"],
        parse_success_rate=processing["success_rate"],
        average_parse_ms=processing["average_parse_ms"],
        shortlist_rate=round(100 * shortlisted / total_candidates, 2) if total_candidates else 0.0,
        rejection_rate=round(100 * rejected / total_candidates, 2) if total_candidates else 0.0,
        average_experience_years=repo.average_experience(session),
        skills_per_candidate=repo.skills_per_candidate(session),
        taxonomy_coverage_percent=repo.taxonomy_coverage(session),
        matches_run=matches_run,
        average_match_score=repo.average_match_score(session),
    )

    status_counts = repo.status_counts(session)
    pipeline = [
        PipelineStage(
            status=status.value,
            label=STATUS_LABELS[status],
            count=status_counts.get(status.value, 0),
            percent=round(100 * status_counts.get(status.value, 0) / total_candidates, 2) if total_candidates else 0.0,
        )
        for status in PIPELINE_ORDER
    ]

    default_gap_skills = gap_skills or [name for name, _, _ in repo.top_skills(session, limit=10)]
    skill_gaps = [item.model_dump() for item in analyze_skill_gaps(session, default_gap_skills)]

    return ReportResponse(
        generated_at=datetime.now(UTC),
        period_start=period_start,
        period_end=period_end,
        kpis=kpis,
        top_technologies=[NamedValue(name=name, value=count) for name, count in repo.technology_distribution(session, limit=12)],
        top_skills=[
            NamedValue(name=name, value=count, extra=category)
            for name, count, category in repo.top_skills(session, limit=15)
        ],
        top_categories=[NamedValue(name=name, value=count) for name, count in repo.category_distribution(session, limit=12)],
        hiring_trends=[
            TrendPoint(period=period, uploads=uploads, candidates=candidates, shortlisted=shortlisted_count, rejected=rejected_count)
            for period, uploads, candidates, shortlisted_count, rejected_count in repo.hiring_trends(session, months=months)
        ],
        skill_gaps=skill_gaps,
        pipeline=pipeline,
        experience_distribution=[
            NamedValue(name=label, value=count) for label, count in repo.experience_distribution(session)
        ],
        top_companies=[NamedValue(name=name, value=count) for name, count in repo.top_companies(session, limit=12)],
        recent_matches=[
            {
                "run_id": run.id,
                "title": run.title,
                "created_at": run.created_at.isoformat(),
                "candidates_evaluated": run.candidates_evaluated,
                "top_score": run.top_score,
                "created_by": run.created_by.full_name if run.created_by else None,
                "top_candidates": [
                    {
                        "candidate_id": result.candidate_id,
                        "name": result.candidate.full_name,
                        "score": result.overall_score,
                        "recommendation": result.recommendation,
                    }
                    for result in run.results[:5]
                ],
            }
            for run in repo.recent_match_runs(session, limit=5)
        ],
    )


# ------------------------------------------------------------------------ exports
def _report_rows(report: ReportResponse) -> list[tuple[str, str, str]]:
    rows: list[tuple[str, str, str]] = [("Section", "Metric", "Value")]
    for field, value in report.kpis.model_dump().items():
        rows.append(("KPI", field.replace("_", " ").title(), str(value)))
    for item in report.top_skills:
        rows.append(("Top Skill", item.name, str(int(item.value))))
    for item in report.top_technologies:
        rows.append(("Top Technology", item.name, str(int(item.value))))
    for item in report.top_categories:
        rows.append(("Skill Category", item.name, str(int(item.value))))
    for stage in report.pipeline:
        rows.append(("Pipeline", stage.label, f"{stage.count} ({stage.percent}%)"))
    for point in report.hiring_trends:
        rows.append(("Hiring Trend", point.period, f"uploads={point.uploads}, candidates={point.candidates}"))
    for gap in report.skill_gaps:
        rows.append(("Skill Gap", str(gap.get("skill")), f"coverage={gap.get('coverage_percent')}%"))
    for item in report.experience_distribution:
        rows.append(("Experience", item.name, str(int(item.value))))
    for item in report.top_companies:
        rows.append(("Company", item.name, str(int(item.value))))
    return rows


def export_csv(report: ReportResponse) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["AI Skill Analyser - Recruitment Report"])
    writer.writerow(["Generated", report.generated_at.isoformat()])
    writer.writerow([])
    for row in _report_rows(report):
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")


def export_excel(report: ReportResponse) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise ValidationAppError("Excel export requires the openpyxl package") from exc

    workbook = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E88E5")

    def write_sheet(title: str, headers: list[str], rows: list[list[object]]) -> None:
        sheet = workbook.create_sheet(title[:31])
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            sheet.append(row)
        for index, header in enumerate(headers, start=1):
            width = max(len(str(header)) + 4, *(len(str(row[index - 1])) + 2 for row in rows)) if rows else len(header) + 4
            sheet.column_dimensions[get_column_letter(index)].width = min(60, width)
        sheet.freeze_panes = "A2"

    workbook.remove(workbook.active)
    write_sheet(
        "KPIs",
        ["Metric", "Value"],
        [[field.replace("_", " ").title(), value] for field, value in report.kpis.model_dump().items()],
    )
    write_sheet(
        "Top Skills",
        ["Skill", "Candidates", "Category"],
        [[item.name, int(item.value), item.extra or ""] for item in report.top_skills],
    )
    write_sheet(
        "Technologies",
        ["Technology Stack", "Candidates"],
        [[item.name, int(item.value)] for item in report.top_technologies],
    )
    write_sheet(
        "Pipeline",
        ["Stage", "Count", "Percent"],
        [[stage.label, stage.count, stage.percent] for stage in report.pipeline],
    )
    write_sheet(
        "Hiring Trends",
        ["Period", "Uploads", "Candidates", "Shortlisted", "Rejected"],
        [
            [point.period, point.uploads, point.candidates, point.shortlisted, point.rejected]
            for point in report.hiring_trends
        ],
    )
    write_sheet(
        "Skill Gaps",
        ["Skill", "Category", "Candidates With Skill", "Coverage %", "Demand Score"],
        [
            [
                gap.get("skill"),
                gap.get("category") or "",
                gap.get("candidates_with_skill"),
                gap.get("coverage_percent"),
                gap.get("demand_score"),
            ]
            for gap in report.skill_gaps
        ],
    )
    write_sheet(
        "Experience",
        ["Band", "Candidates"],
        [[item.name, int(item.value)] for item in report.experience_distribution],
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_pdf(report: ReportResponse) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:  # pragma: no cover
        raise ValidationAppError("PDF export requires the reportlab package") from exc

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        title="AI Skill Analyser - Recruitment Report",
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    heading = ParagraphStyle("heading", parent=styles["Heading2"], textColor=colors.HexColor("#0D47A1"))
    story: list[object] = [
        Paragraph("AI Skill Analyser", styles["Title"]),
        Paragraph(
            f"Recruitment report generated {report.generated_at:%d %b %Y %H:%M UTC}",
            styles["Normal"],
        ),
        Spacer(1, 8 * mm),
    ]

    def add_table(title: str, headers: list[str], rows: list[list[object]]) -> None:
        if not rows:
            return
        story.append(Paragraph(title, heading))
        story.append(Spacer(1, 3 * mm))
        table = Table([headers, *rows], hAlign="LEFT", repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E88E5")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BBDEFB")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F9FF")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.extend([table, Spacer(1, 7 * mm)])

    add_table(
        "Recruitment KPIs",
        ["Metric", "Value"],
        [[field.replace("_", " ").title(), str(value)] for field, value in report.kpis.model_dump().items()],
    )
    add_table(
        "Top Skills",
        ["Skill", "Candidates", "Category"],
        [[item.name, int(item.value), item.extra or "-"] for item in report.top_skills[:15]],
    )
    add_table(
        "Top Technologies",
        ["Technology", "Candidates"],
        [[item.name, int(item.value)] for item in report.top_technologies[:12]],
    )
    add_table(
        "Candidate Pipeline",
        ["Stage", "Count", "Share"],
        [[stage.label, stage.count, f"{stage.percent}%"] for stage in report.pipeline],
    )
    add_table(
        "Hiring Trends",
        ["Period", "Uploads", "New Candidates", "Shortlisted", "Rejected"],
        [
            [point.period, point.uploads, point.candidates, point.shortlisted, point.rejected]
            for point in report.hiring_trends
        ],
    )
    add_table(
        "Skill Gap Analysis",
        ["Skill", "Coverage %", "Candidates", "Suggested bridge skills"],
        [
            [
                gap.get("skill"),
                gap.get("coverage_percent"),
                gap.get("candidates_with_skill"),
                ", ".join(gap.get("suggested_learning") or [])[:60] or "-",
            ]
            for gap in report.skill_gaps[:15]
        ],
    )

    document.build(story)
    return buffer.getvalue()


EXPORT_FORMATS = {
    "csv": ("text/csv", "csv", export_csv),
    "excel": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx", export_excel),
    "xlsx": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx", export_excel),
    "pdf": ("application/pdf", "pdf", export_pdf),
}


def export_report(report: ReportResponse, export_format: str) -> tuple[bytes, str, str]:
    key = (export_format or "").lower()
    if key not in EXPORT_FORMATS:
        raise ValidationAppError(f"Unsupported export format '{export_format}'. Use csv, excel or pdf.")
    media_type, extension, exporter = EXPORT_FORMATS[key]
    payload = exporter(report)
    filename = f"skill-analyser-report-{report.generated_at:%Y%m%d-%H%M}.{extension}"
    return payload, media_type, filename


def export_candidates_csv(session: Session, candidate_ids: list[int] | None = None) -> bytes:
    from app.schemas.candidate import CandidateFilters

    filters = CandidateFilters(page=1, page_size=1000, candidate_ids=candidate_ids)
    candidates, _ = repo.list_candidates(session, filters)

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "Candidate ID",
            "Name",
            "Email",
            "Phone",
            "Current Title",
            "Current Company",
            "Location",
            "Experience (yrs)",
            "Highest Degree",
            "Status",
            "AI Score",
            "Top Skills",
            "Created At",
        ]
    )
    for candidate in candidates:
        writer.writerow(
            [
                candidate.id,
                candidate.full_name,
                candidate.email or "",
                candidate.phone or "",
                candidate.current_title or "",
                candidate.current_company_name or "",
                ", ".join(part for part in [candidate.city, candidate.country] if part),
                candidate.total_experience_years,
                candidate.highest_degree or "",
                candidate.status,
                candidate.last_match_score if candidate.last_match_score is not None else "",
                "; ".join(link.display_name for link in candidate.skills[:12]),
                candidate.created_at.isoformat(),
            ]
        )
    return buffer.getvalue().encode("utf-8-sig")
